"""Tradebook xlsx importer.

Parses a Zerodha tradebook .xlsx (downloaded from Console) and uses FIFO
logic per symbol to determine the entry_date of currently-held quantity.

The tradebook layout is:
  - Rows 1..14 = header / metadata (skipped)
  - Row 15 = column headers
  - Row 16+ = trades, one row per fill, sorted by trade_date asc

We compute, for each symbol, the date at which the *current* held quantity
was acquired. Algorithm:
  - Build chronological list of (date, side, qty) per symbol from the file
  - Walk forward: BUY adds to a queue, SELL pops from oldest (FIFO)
  - The entry_date for the remaining queue is min(remaining queue's dates)

Result is upserted into falcon_position_first_seen with detected_via='tradebook'.
Existing rows for symbols with detected_via='first_poll' or 'kite_t1' are
overwritten — tradebook is more authoritative.
"""
from __future__ import annotations

import io
import logging
from collections import defaultdict, deque
from datetime import datetime, date
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from ...db import falcon_conn

log = logging.getLogger("kanida.falcon.trade.tradebook_import")

# Zerodha tradebook layout is stable; column header row is 15 (1-indexed).
HEADER_ROW = 15
DATA_START_ROW = 16


def _parse_xlsx(data: bytes) -> List[Dict[str, Any]]:
    """Extract trade rows from an in-memory xlsx. Returns list of dicts with
    keys: symbol, side ('BUY'/'SELL'), qty (int), trade_date (date)."""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []

    # Find column indices from the header row.
    header_values: Dict[str, int] = {}
    for cell in ws[HEADER_ROW]:
        v = cell.value
        if isinstance(v, str):
            header_values[v.strip().lower()] = cell.column - 1  # 0-indexed

    # Defensive — Zerodha column names have varied slightly across years.
    def _find(*names: str) -> Optional[int]:
        for n in names:
            if n in header_values:
                return header_values[n]
        return None

    col_symbol = _find("symbol", "tradingsymbol", "instrument")
    col_side   = _find("trade_type", "side", "transaction_type", "buy/sell")
    col_qty    = _find("quantity", "qty")
    col_date   = _find("trade_date", "order_execution_time", "trade date", "trade time")

    if any(c is None for c in (col_symbol, col_side, col_qty, col_date)):
        raise ValueError(
            f"Tradebook xlsx missing expected columns. Found header: {list(header_values.keys())}. "
            "Need symbol, side, qty, trade_date."
        )

    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        try:
            sym = row[col_symbol]
            side = row[col_side]
            qty = row[col_qty]
            td = row[col_date]
            if not sym or qty is None or td is None or not side:
                continue
            sym = str(sym).strip()
            side_norm = str(side).strip().upper()
            if side_norm not in ("BUY", "SELL"):
                continue
            qty_int = int(qty)
            if qty_int <= 0:
                continue
            # Date column may be datetime, date, or string
            if isinstance(td, datetime):
                trade_date = td.date()
            elif isinstance(td, date):
                trade_date = td
            else:
                # try common formats
                td_str = str(td).strip()
                trade_date = None
                for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d-%m-%Y", "%d/%m/%Y"):
                    try:
                        trade_date = datetime.strptime(td_str, fmt).date()
                        break
                    except ValueError:
                        continue
                if trade_date is None:
                    continue
            rows.append({
                "symbol":     sym,
                "side":       side_norm,
                "qty":        qty_int,
                "trade_date": trade_date,
            })
        except Exception as e:
            log.warning("tradebook row parse failed (skipping): %s", e)
            continue
    return rows


def _fifo_entry_dates(rows: List[Dict[str, Any]]) -> Dict[str, Tuple[date, int]]:
    """Per symbol, walk trades chronologically. BUYs append (date, qty) to a
    FIFO queue; SELLs pop from the oldest. Returns {symbol: (entry_date, qty)}
    where entry_date = oldest date in remaining queue, qty = remaining qty.
    Symbols with zero remaining quantity are excluded."""
    by_symbol: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in rows:
        by_symbol[r["symbol"]].append(r)

    out: Dict[str, Tuple[date, int]] = {}
    for sym, srows in by_symbol.items():
        # Sort chronologically — same-day trades fall in the order of the file.
        srows.sort(key=lambda r: (r["trade_date"], 0 if r["side"] == "BUY" else 1))
        queue: deque[Tuple[date, int]] = deque()
        for r in srows:
            if r["side"] == "BUY":
                queue.append((r["trade_date"], r["qty"]))
            else:
                # SELL — pop from oldest until qty satisfied
                remaining = r["qty"]
                while remaining > 0 and queue:
                    head_date, head_qty = queue[0]
                    if head_qty > remaining:
                        queue[0] = (head_date, head_qty - remaining)
                        remaining = 0
                    else:
                        queue.popleft()
                        remaining -= head_qty
        if not queue:
            continue
        oldest_date = queue[0][0]
        total_qty = sum(q for _, q in queue)
        out[sym] = (oldest_date, total_qty)
    return out


def import_xlsx(data: bytes) -> Dict[str, Any]:
    """Parse + FIFO-derive entry dates + upsert into falcon_position_first_seen.
    Returns summary."""
    trades = _parse_xlsx(data)
    if not trades:
        return {"status": "empty", "trades_parsed": 0, "symbols_dated": 0}

    derived = _fifo_entry_dates(trades)
    if not derived:
        return {"status": "no_open", "trades_parsed": len(trades), "symbols_dated": 0}

    now_iso = datetime.now().isoformat()
    inserted = 0
    updated = 0
    with falcon_conn() as con:
        for sym, (entry_date, qty) in derived.items():
            row = con.execute(
                "SELECT detected_via FROM falcon_position_first_seen WHERE symbol=?",
                (sym,),
            ).fetchone()
            if row is None:
                con.execute("""
                    INSERT INTO falcon_position_first_seen
                      (symbol, first_seen_date, first_seen_at, detected_via,
                       qty_at_first_seen)
                    VALUES (?,?,?,?,?)
                """, (sym, entry_date.isoformat(), now_iso, "tradebook", qty))
                inserted += 1
            else:
                # Tradebook overwrites first_poll/kite_t1 (less authoritative).
                # Don't overwrite an earlier 'tradebook' row — file may have
                # duplicates in a re-import case; trust the first import's date.
                if (row["detected_via"] or "") != "tradebook":
                    con.execute("""
                        UPDATE falcon_position_first_seen
                        SET first_seen_date = ?, first_seen_at = ?,
                            detected_via = ?, qty_at_first_seen = ?
                        WHERE symbol = ?
                    """, (entry_date.isoformat(), now_iso, "tradebook", qty, sym))
                    updated += 1
        con.commit()

    return {
        "status":         "ok",
        "trades_parsed":  len(trades),
        "symbols_dated":  len(derived),
        "inserted":       inserted,
        "updated":        updated,
        "symbols":        sorted(derived.keys()),
    }
