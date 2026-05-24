"""DRAFT — Parity test: backend numbers MUST match the desktop audit Excels.

This is the safety net. If the refactor is done correctly, every yearly
return, win rate, max DD, and trade count produced by
`persona_simulator.simulate_persona(slug)` MUST exactly match the same
field in the audit Excel `persona{N}_V3_*.xlsx` files on operator's Desktop.

If parity fails, the refactor is broken and we must NOT ship.

USAGE:
    pytest backend/power_user/tests/test_persona_parity.py -v

Tolerance:
  - Return % / WR % / DD %: 0.05 pp absolute (rounding-safe)
  - Trade counts: exact match (zero tolerance)
  - P&L Rs: <= Re 1.00 absolute (yearly reconciliation tolerance)
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional

import pytest

from ..services.persona_simulator import PERSONA_CONFIGS, simulate_persona


# ════════════════════════════════════════════════════════════════════════
# Excel paths — these are the canonical lock files on operator's Desktop
# ════════════════════════════════════════════════════════════════════════

DESKTOP = Path.home() / "Desktop"

EXCEL_BY_SLUG = {
    "daily-trader":   DESKTOP / "persona1_V3_fixed35k_2021_2026.xlsx",
    "patient-trader": DESKTOP / "persona2_V3_fixed35k_2024_2026.xlsx",
    "weekly-trader":  DESKTOP / "persona3_V3_variantA2_fixed_50k_2021_2026.xlsx",
    "monthly-trader": DESKTOP / "persona4_V3_locked_top14_sl10_tgt30.xlsx",
    "btst-trader":    DESKTOP / "persona5_V3_locked_top15_sl5_tgt7.xlsx",
}


def _read_excel_yearly(path: Path) -> List[Dict[str, Any]]:
    """Read the 'Summary' sheet from an audit Excel and return year rows."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb["Summary"]
    # Header row is row 4 in all the audit Excels (see write_audit_excel format)
    headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    rows = []
    r = 5
    while True:
        year_val = ws.cell(row=r, column=1).value
        # Stop when we hit AGGREGATE row or blank
        if year_val is None or not isinstance(year_val, int):
            break
        row = {}
        for ci, h in enumerate(headers, start=1):
            row[h] = ws.cell(row=r, column=ci).value
        rows.append(row)
        r += 1
    return rows


def _norm_yearly_row(excel_row: Dict[str, Any]) -> Dict[str, Any]:
    """Normalise Excel row keys → our backend yearly dict keys."""
    return {
        "year": excel_row.get("Year"),
        "return_pct": excel_row.get("Return %"),
        "max_dd_pct": excel_row.get("Max DD %"),
        "win_rate_pct": excel_row.get("Win Rate %"),
        "n_closed": excel_row.get("Closed"),
        "n_open_at_year_end": excel_row.get("Open at EoY") or excel_row.get("Open EoY"),
        "end_equity": excel_row.get("End Equity"),
        "pnl": excel_row.get("P&L"),
    }


# ════════════════════════════════════════════════════════════════════════
# Pytest parametrised parity test
# ════════════════════════════════════════════════════════════════════════

@pytest.mark.parametrize("slug", list(PERSONA_CONFIGS.keys()))
def test_yearly_table_matches_excel(slug):
    """Every year row from the backend MUST match the audit Excel."""
    excel_path = EXCEL_BY_SLUG.get(slug)
    if not excel_path or not excel_path.exists():
        pytest.skip(f"Audit Excel for {slug} not found at {excel_path}. "
                    f"Cannot run parity test without it.")

    excel_rows = [_norm_yearly_row(r) for r in _read_excel_yearly(excel_path)]
    backend_result = simulate_persona(slug, force=True)
    backend_rows = backend_result["yearly"]

    assert len(backend_rows) == len(excel_rows), (
        f"{slug}: backend has {len(backend_rows)} years, Excel has {len(excel_rows)}"
    )

    for exc_row, be_row in zip(excel_rows, backend_rows):
        Y = exc_row["year"]
        # Trade count: zero tolerance
        assert be_row["n_closed"] == exc_row["n_closed"], (
            f"{slug} {Y}: closed trades mismatch — backend {be_row['n_closed']} vs Excel {exc_row['n_closed']}"
        )
        # Return % within 0.05 pp
        assert abs(be_row["return_pct"] - exc_row["return_pct"]) < 0.05, (
            f"{slug} {Y}: return mismatch — backend {be_row['return_pct']:.2f}% vs Excel {exc_row['return_pct']:.2f}%"
        )
        # Win rate within 0.05 pp
        assert abs(be_row["win_rate_pct"] - exc_row["win_rate_pct"]) < 0.05, (
            f"{slug} {Y}: WR mismatch — backend {be_row['win_rate_pct']:.2f}% vs Excel {exc_row['win_rate_pct']:.2f}%"
        )
        # Max DD within 0.05 pp
        assert abs(be_row["max_dd_pct"] - exc_row["max_dd_pct"]) < 0.05, (
            f"{slug} {Y}: max DD mismatch — backend {be_row['max_dd_pct']:.2f}% vs Excel {exc_row['max_dd_pct']:.2f}%"
        )
        # End equity within Re 1.00
        assert abs(be_row["end_equity"] - exc_row["end_equity"]) < 1.00, (
            f"{slug} {Y}: end equity mismatch — backend ₹{be_row['end_equity']:,.2f} vs Excel ₹{exc_row['end_equity']:,.2f}"
        )


@pytest.mark.parametrize("slug", list(PERSONA_CONFIGS.keys()))
def test_reconciliation_passes(slug):
    """P&L must reconcile year-by-year to <Re 1.00 gap."""
    excel_path = EXCEL_BY_SLUG.get(slug)
    if not excel_path or not excel_path.exists():
        pytest.skip(f"Audit Excel for {slug} not found")

    result = simulate_persona(slug, force=True)
    for row in result["reconciliation"]:
        assert abs(row["gap"]) < 1.00, (
            f"{slug} {row['year']}: reconciliation gap ₹{row['gap']:,.2f} exceeds Re 1.00"
        )


def test_all_personas_have_required_fields():
    """API contract test — every persona returns the expected shape."""
    for slug in PERSONA_CONFIGS.keys():
        excel_path = EXCEL_BY_SLUG.get(slug)
        if not excel_path or not excel_path.exists():
            continue
        result = simulate_persona(slug, force=True)
        # Top-level keys
        for key in ("persona_id", "name", "tagline", "risk_tier", "config",
                     "yearly", "monthly", "trades", "summary", "reconciliation",
                     "open_positions", "last_updated", "source_sim_script"):
            assert key in result, f"{slug}: missing key '{key}'"
        # Summary keys
        for key in ("avg_yearly_return_pct", "median_yearly_return_pct",
                     "best_year_pct", "worst_year_pct", "positive_years",
                     "total_years", "avg_max_drawdown_pct", "avg_win_rate_pct"):
            assert key in result["summary"], f"{slug}: summary missing '{key}'"
        # Yearly row keys
        if result["yearly"]:
            for key in ("year", "return_pct", "max_dd_pct", "win_rate_pct",
                         "n_closed", "n_open_at_year_end", "end_equity", "pnl"):
                assert key in result["yearly"][0], f"{slug}: yearly row missing '{key}'"


def test_cache_works():
    """Second call should be instant (cache hit)."""
    import time
    slug = "monthly-trader"  # smallest sim, fastest baseline
    if not EXCEL_BY_SLUG[slug].exists():
        pytest.skip()
    # Cold run
    simulate_persona(slug, force=True)
    # Warm run
    t0 = time.time()
    simulate_persona(slug, force=False)
    elapsed = time.time() - t0
    assert elapsed < 0.1, f"Cache hit should be <100ms, got {elapsed*1000:.0f}ms"
