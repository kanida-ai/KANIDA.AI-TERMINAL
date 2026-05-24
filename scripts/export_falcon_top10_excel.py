"""Export the live /power/today Falcon Top 10 to Excel — ONE SHEET PER STOCK,
laid out vertically to mirror the card UX (not a database dump).

For each pick, the sheet renders top-to-bottom:

  HEADER STRIP        rank · symbol · sector · badges (signal type, risk,
                      today HR, FRESH/CAP flags) · "N fires · combined +X pp"

  ① WHY IS THE AI PICKING THIS?            (right-aligned tag)
     [synthesis paragraph — green left border]
     CONFLUENCE BREAKDOWN · N regimes firing
     [regime row 1]
     [regime row 2]
     ...

  ② HISTORICAL EVIDENCE                    (right-aligned tag)
     [HIT RATE]  [AVG WIN]  [AVG LOSS]  [WORST DRAWDOWN]
     [green callout — today_HR vs lifetime baseline + last similar]

  ③ SECTOR CONTEXT                         (right-aligned: TAILWIND/etc)
     [SECTOR RANK]  [ROTATION]  [SECTOR 20D]  [PEERS IN TOP 50]
     [sector narrative paragraph]

  ACTION  |  STOP LOSS  |  TIME HORIZON

Plus an "Index" sheet at the front linking to each per-stock sheet.

Output: C:\\Users\\SPS\\Desktop\\falcon_top10_ux_export.xlsx
"""
from __future__ import annotations

import json
import sys
import urllib.request
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.hyperlink import Hyperlink
except ImportError:
    sys.stderr.write("ERROR: openpyxl not installed. pip install openpyxl\n")
    sys.exit(1)

ENDPOINT = "http://127.0.0.1:8001/api/power/today/falcon-top-20"
OUT_PATH = Path(r"C:\Users\SPS\Desktop\falcon_top10_ux_export.xlsx")


# ── Card palette (mirrors the dark Tailwind theme on /power/today) ──────
COL_BG_CARD     = "0A0A0A"   # near-black card background
COL_BG_SECTION  = "171717"   # slightly lighter section dividers
COL_BG_STAT     = "171717"   # stat box bg
COL_BG_CALLOUT  = "0E2A1F"   # green-tinted callout bg
COL_BG_ACTION   = "0E2A1F"   # green-tinted action box bg
COL_BORDER      = "262626"
COL_BORDER_MINT = "1F5A3C"
COL_MINT_400    = "3FE3A4"
COL_MINT_300    = "5AECB5"
COL_AMBER_300   = "FCD34D"
COL_RED_300     = "FCA5A5"
COL_BLUE_300    = "93C5FD"
COL_WHITE       = "FFFFFF"
COL_WHITE_85    = "D4D4D4"
COL_WHITE_55    = "8E8E8E"
COL_WHITE_45    = "747474"

# ── Reusable styles ─────────────────────────────────────────────────────
def F(size=10, bold=False, color=COL_WHITE_85, name="Calibri", italic=False):
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

ALN_L  = Alignment(horizontal="left",   vertical="center", wrap_text=False)
ALN_LW = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
ALN_C  = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALN_R  = Alignment(horizontal="right",  vertical="center", wrap_text=False)

def FILL(hex6: str) -> PatternFill:
    return PatternFill("solid", start_color=hex6, end_color=hex6)

# Thin border for stat boxes
def BORDER(color=COL_BORDER) -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

# Left-only border (for "left-border accent" on synthesis box)
def LEFTBAR(color=COL_MINT_400) -> Border:
    return Border(left=Side(style="thick", color=color))


# ── HTTP fetch ──────────────────────────────────────────────────────────
def fetch_payload() -> dict:
    print(f"GET {ENDPOINT} ...")
    with urllib.request.urlopen(ENDPOINT, timeout=60) as r:
        body = r.read().decode("utf-8")
    payload = json.loads(body)
    print(f"  signal_date={payload.get('signal_date')}  n_picks={len(payload.get('picks', []))}")
    return payload


# ── Per-stock sheet writer ──────────────────────────────────────────────
def safe_sheet_name(rank: int, sym: str) -> str:
    raw = f"{rank:02d} {sym}"
    # Excel max 31 chars, no /\?*[]:
    bad = "/\\?*[]:"
    for ch in bad:
        raw = raw.replace(ch, "")
    return raw[:31]


def write_stock_sheet(wb: Workbook, pick: dict, signal_date: str, entry_date: str) -> str:
    name = safe_sheet_name(pick.get("rank", 0), pick.get("symbol", "?"))
    ws = wb.create_sheet(name)

    # Page is 8 cols wide. Stat boxes = 2 cols each × 3 rows tall.
    # Column widths to roughly match the visual rhythm of the card.
    widths = {"A": 14, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Whole sheet background = dark card
    def paint_row(row_idx: int, fill_hex: str = COL_BG_CARD, height: int = 18):
        for col in range(1, 9):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = FILL(fill_hex)
        ws.row_dimensions[row_idx].height = height

    # ── 1. HEADER STRIP ─────────────────────────────────────────────────
    b1 = pick.get("bucket1", {}) or {}
    b2 = pick.get("bucket2", {}) or {}
    b3 = pick.get("bucket3", {}) or {}
    flags = pick.get("flags", {}) or {}
    action = pick.get("action", {}) or {}

    paint_row(1, COL_BG_CARD, height=30)
    title_cell = ws.cell(row=1, column=1,
                          value=f"#{pick['rank']}  {pick['symbol']}  ·  {pick.get('sector') or '—'}")
    title_cell.font = F(size=16, bold=True, color=COL_WHITE)
    title_cell.alignment = ALN_L
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=8)

    # Badge row
    paint_row(2, COL_BG_CARD, height=22)
    badge_text = []
    badge_text.append(f"★ {pick.get('signal_type', '').upper()}")
    badge_text.append(f"{pick.get('risk_level', '').upper()} RISK")
    if b1.get("today_weighted_hit_rate") is not None:
        badge_text.append(f"TODAY HR {round(b1['today_weighted_hit_rate'])}%")
    if flags.get("fresh_entry"):           badge_text.append("✨ FRESH")
    if flags.get("falling_knife"):         badge_text.append("⚠ FALLING KNIFE")
    elif flags.get("capitulation_warning"): badge_text.append("⚠ CAPITULATION")
    ws.cell(row=2, column=1, value="   ".join(badge_text)).font = F(
        size=10, bold=True, color=COL_MINT_300)
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=5)
    summary_right = (f"{b1.get('pattern_count', 0)} fires  ·  "
                     f"combined +{b1.get('combined_lift_pp', 0):.0f}pp")
    rc = ws.cell(row=2, column=6, value=summary_right)
    rc.font = F(size=10, color=COL_WHITE_55); rc.alignment = ALN_R
    ws.merge_cells(start_row=2, end_row=2, start_column=6, end_column=8)

    # ── 2. BUCKET 1 — WHY ───────────────────────────────────────────────
    paint_row(3, COL_BG_CARD, height=6)   # spacer

    paint_row(4, COL_BG_SECTION, height=24)
    h1 = ws.cell(row=4, column=1,
                  value=f"①  WHY IS THE AI PICKING THIS?")
    h1.font = F(size=12, bold=True, color=COL_MINT_400)
    ws.merge_cells(start_row=4, end_row=4, start_column=1, end_column=5)
    h1tag = ws.cell(row=4, column=6,
                     value=f"{b1.get('pattern_count', 0)} patterns fired  ·  combined +{b1.get('combined_lift_pp', 0):.0f}pp lift")
    h1tag.font = F(size=10, color=COL_WHITE_45); h1tag.alignment = ALN_R
    ws.merge_cells(start_row=4, end_row=4, start_column=6, end_column=8)

    # Synthesis box — merged 3 rows tall, wraps
    synth = b1.get("synthesis", "—")
    paint_row(5, COL_BG_CARD, height=22)
    paint_row(6, COL_BG_CARD, height=22)
    paint_row(7, COL_BG_CARD, height=22)
    sc = ws.cell(row=5, column=1, value=synth)
    sc.font = F(size=11, color=COL_WHITE_85); sc.alignment = ALN_LW
    sc.border = LEFTBAR(COL_MINT_400)
    ws.merge_cells(start_row=5, end_row=7, start_column=1, end_column=8)

    # CONFLUENCE BREAKDOWN header
    paint_row(8, COL_BG_CARD, height=8)
    breakdown = b1.get("regime_breakdown") or []
    paint_row(9, COL_BG_CARD, height=20)
    cb = ws.cell(row=9, column=1,
                  value=f"CONFLUENCE BREAKDOWN  ·  {len(breakdown)} regime{'s' if len(breakdown)!=1 else ''} firing")
    cb.font = F(size=9, bold=True, color=COL_WHITE_55)
    ws.merge_cells(start_row=9, end_row=9, start_column=1, end_column=8)

    # One row per regime — 4-col layout: [regime+headline] [count/pct] [avg lift] [HR]
    total_fires = b1.get("pattern_count") or 1
    cur_row = 10
    for g in breakdown:
        regime = (g.get("regime") or "other").lower()
        glyph = REGIME_GLYPH.get(regime, "·")
        regime_pretty = REGIME_HUMAN.get(regime, regime.title())
        accent = REGIME_COLOR.get(regime, COL_WHITE_85)
        headline = g.get("headline") or ""
        count = g.get("count", 0)
        pct = (count / total_fires * 100) if total_fires else 0
        avg_lift = g.get("avg_lift_pp", 0)
        hr = g.get("weighted_hit_rate")
        if hr is None: hr = g.get("avg_oos_hit_rate")

        paint_row(cur_row,     COL_BG_STAT, height=18)
        paint_row(cur_row + 1, COL_BG_STAT, height=14)
        ws.row_dimensions[cur_row + 1].height = 14

        # Regime + headline (cols A-D)
        rcell = ws.cell(row=cur_row, column=1, value=f"{glyph}  {regime_pretty}")
        rcell.font = F(size=11, bold=True, color=accent)
        rcell.alignment = ALN_L
        for c in range(1, 5):
            ws.cell(row=cur_row, column=c).fill = FILL(COL_BG_STAT)
            ws.cell(row=cur_row, column=c).border = BORDER()
        ws.merge_cells(start_row=cur_row, end_row=cur_row, start_column=2, end_column=4)
        ws.cell(row=cur_row, column=2, value=headline).font = F(size=9, color=COL_WHITE_55)
        ws.cell(row=cur_row, column=2).alignment = ALN_L

        # Sub-line: count / pct
        ws.cell(row=cur_row + 1, column=1,
                 value=f"   {count} of {total_fires} patterns ({pct:.0f}%)").font = F(
            size=9, color=COL_WHITE_55)
        for c in range(1, 5):
            ws.cell(row=cur_row + 1, column=c).fill = FILL(COL_BG_STAT)
            ws.cell(row=cur_row + 1, column=c).border = BORDER()
        ws.merge_cells(start_row=cur_row + 1, end_row=cur_row + 1, start_column=1, end_column=4)

        # Avg lift (cols E-F)
        for c in range(5, 7):
            ws.cell(row=cur_row, column=c).fill = FILL(COL_BG_STAT)
            ws.cell(row=cur_row, column=c).border = BORDER()
            ws.cell(row=cur_row + 1, column=c).fill = FILL(COL_BG_STAT)
            ws.cell(row=cur_row + 1, column=c).border = BORDER()
        ws.merge_cells(start_row=cur_row, end_row=cur_row, start_column=5, end_column=6)
        ws.merge_cells(start_row=cur_row + 1, end_row=cur_row + 1, start_column=5, end_column=6)
        lift_cell = ws.cell(row=cur_row, column=5, value=f"+{avg_lift:.1f}pp")
        lift_cell.font = F(size=14, bold=True, color=COL_MINT_400, name="Consolas")
        lift_cell.alignment = ALN_R
        ws.cell(row=cur_row + 1, column=5, value="avg lift").font = F(size=8, color=COL_WHITE_45)
        ws.cell(row=cur_row + 1, column=5).alignment = ALN_R

        # HR (cols G-H)
        for c in range(7, 9):
            ws.cell(row=cur_row, column=c).fill = FILL(COL_BG_STAT)
            ws.cell(row=cur_row, column=c).border = BORDER()
            ws.cell(row=cur_row + 1, column=c).fill = FILL(COL_BG_STAT)
            ws.cell(row=cur_row + 1, column=c).border = BORDER()
        ws.merge_cells(start_row=cur_row, end_row=cur_row, start_column=7, end_column=8)
        ws.merge_cells(start_row=cur_row + 1, end_row=cur_row + 1, start_column=7, end_column=8)
        hr_text = f"{round(hr)}%" if hr is not None else "—"
        hrc = ws.cell(row=cur_row, column=7, value=hr_text)
        hrc.font = F(size=14, bold=True, color=COL_MINT_300, name="Consolas")
        hrc.alignment = ALN_R
        ws.cell(row=cur_row + 1, column=7, value="weighted HR").font = F(size=8, color=COL_WHITE_45)
        ws.cell(row=cur_row + 1, column=7).alignment = ALN_R

        cur_row += 2

    # ── 3. BUCKET 2 — HISTORICAL EVIDENCE ────────────────────────────────
    cur_row += 1
    paint_row(cur_row, COL_BG_CARD, height=8)
    cur_row += 1

    paint_row(cur_row, COL_BG_SECTION, height=24)
    h2 = ws.cell(row=cur_row, column=1, value="②  HISTORICAL EVIDENCE")
    h2.font = F(size=12, bold=True, color=COL_MINT_400)
    ws.merge_cells(start_row=cur_row, end_row=cur_row, start_column=1, end_column=5)
    nfires = b2.get("n_historical_fires", 0)
    nout   = b2.get("n_with_outcome", 0)
    lb     = b2.get("lookback_start", "2021-01-01")
    tag2 = (f"{nout:,} of {nfires:,} pattern fires realised (from {lb})"
            if nout > 0 else
            f"{nfires:,} pattern fires (from {lb}) · outcomes still pending")
    h2tag = ws.cell(row=cur_row, column=6, value=tag2)
    h2tag.font = F(size=10, color=COL_WHITE_45); h2tag.alignment = ALN_R
    ws.merge_cells(start_row=cur_row, end_row=cur_row, start_column=6, end_column=8)
    cur_row += 1

    paint_row(cur_row, COL_BG_CARD, height=8)
    cur_row += 1

    # 4 stat boxes side-by-side: HIT RATE | AVG WIN | AVG LOSS | WORST DD
    has_outcomes = nout > 0
    stats = [
        ("HIT RATE" if has_outcomes else "LIFETIME HIT RATE",
            f"{b2.get('hit_rate_pct', 0):.0f}%",
            ("of fires hit +10% in 20 days" if has_outcomes
                else "stock's all-time +10%/20d rate"),
            COL_MINT_400),
        ("AVG WIN",
            (f"+{b2['avg_win_pct']:.1f}%" if b2.get("avg_win_pct") is not None else "—"),
            "when the pattern works",
            COL_MINT_400 if b2.get("avg_win_pct") is not None else COL_WHITE_45),
        ("AVG LOSS",
            (f"{b2['avg_loss_pct']:.1f}%" if b2.get("avg_loss_pct") is not None else "—"),
            "when the pattern fails",
            COL_RED_300 if b2.get("avg_loss_pct") is not None else COL_WHITE_45),
        ("WORST DRAWDOWN",
            (f"{b2['worst_drawdown_pct']:.1f}%" if b2.get("worst_drawdown_pct") is not None
                else f"{b2.get('lifetime_baseline_pct', 0):.0f}% (baseline)"),
            "worst MAE within 20d" if b2.get("worst_drawdown_pct") is not None
                else "all-time +10%/20d rate",
            COL_RED_300 if b2.get("worst_drawdown_pct") is not None else COL_WHITE_45),
    ]
    _draw_stat_row(ws, cur_row, stats)
    cur_row += 3

    # Green callout
    callout = _bucket2_callout_text(b2)
    paint_row(cur_row, COL_BG_CALLOUT, height=22)
    paint_row(cur_row+1, COL_BG_CALLOUT, height=22)
    paint_row(cur_row+2, COL_BG_CALLOUT, height=22)
    cc = ws.cell(row=cur_row, column=1, value=callout)
    cc.font = F(size=10, color=COL_MINT_300); cc.alignment = ALN_LW
    cc.border = BORDER(COL_BORDER_MINT)
    for c in range(1, 9):
        ws.cell(row=cur_row+1, column=c).fill = FILL(COL_BG_CALLOUT)
        ws.cell(row=cur_row+2, column=c).fill = FILL(COL_BG_CALLOUT)
    ws.merge_cells(start_row=cur_row, end_row=cur_row+2, start_column=1, end_column=8)
    cur_row += 3

    # ── 4. BUCKET 3 — SECTOR CONTEXT ─────────────────────────────────────
    cur_row += 1
    paint_row(cur_row, COL_BG_CARD, height=8)
    cur_row += 1

    paint_row(cur_row, COL_BG_SECTION, height=24)
    h3 = ws.cell(row=cur_row, column=1, value="③  SECTOR CONTEXT")
    h3.font = F(size=12, bold=True, color=COL_MINT_400)
    ws.merge_cells(start_row=cur_row, end_row=cur_row, start_column=1, end_column=5)
    verdict_glyph = {"tailwind": "↑ TAILWIND", "headwind": "↓ HEADWIND",
                      "neutral":  "→ NEUTRAL",  "unranked": "UNRANKED"}.get(
                          b3.get("verdict"), b3.get("verdict", "").upper())
    h3tag = ws.cell(row=cur_row, column=6,
                     value=f"{b3.get('sector_name', '—')} sector  ·  {verdict_glyph}")
    h3tag.font = F(size=10, color=_verdict_color(b3.get("verdict")))
    h3tag.alignment = ALN_R
    ws.merge_cells(start_row=cur_row, end_row=cur_row, start_column=6, end_column=8)
    cur_row += 1

    paint_row(cur_row, COL_BG_CARD, height=8)
    cur_row += 1

    # 4 sector stat boxes
    rank_now = b3.get("sector_rank", 0)
    rank_5d  = b3.get("sector_rank_5d_ago")
    total    = b3.get("total_sectors", 0)
    rot_pos  = b3.get("rotation_positions", 0)
    rot_dir  = b3.get("rotation_direction", "")
    rot_label = (f"↑ +{rot_pos}" if rot_pos > 0
                  else f"↓ {rot_pos}" if rot_pos < 0
                  else "flat")
    sector_stats = [
        ("SECTOR RANK", f"#{rank_now}",
            (f"of {total} · was #{rank_5d} 5d ago" if rank_5d else f"of {total} sectors today"),
            COL_MINT_400 if rank_now and rank_now <= 7 else (COL_RED_300 if rank_now >= 14 else COL_WHITE_85)),
        ("ROTATION", rot_label,
            ("rotating IN (5d)" if rot_pos > 0 else "rotating OUT (5d)" if rot_pos < 0 else "unchanged (5d)"),
            COL_MINT_400 if rot_pos > 1 else (COL_RED_300 if rot_pos < -1 else COL_WHITE_85)),
        ("SECTOR 20D",
            f"{b3.get('sector_20d_return_pct', 0):+.1f}%",
            "last 20 trading days",
            COL_MINT_400 if b3.get("sector_20d_return_pct", 0) >= 0 else COL_RED_300),
        ("PEERS IN TOP 50", f"{b3.get('peer_count_in_top50', 0)}",
            ("broad-based sector strength" if b3.get("peer_count_in_top50", 0) >= 3
                else "some sector company" if b3.get("peer_count_in_top50", 0) >= 1
                else "idiosyncratic (alone)"),
            COL_MINT_400 if b3.get("peer_count_in_top50", 0) >= 3 else COL_WHITE_85),
    ]
    _draw_stat_row(ws, cur_row, sector_stats)
    cur_row += 3

    # Sector narrative
    narr = _sector_narrative(b3)
    paint_row(cur_row, COL_BG_CARD, height=22)
    paint_row(cur_row+1, COL_BG_CARD, height=22)
    snarr = ws.cell(row=cur_row, column=1, value=narr)
    snarr.font = F(size=10, color=COL_WHITE_85); snarr.alignment = ALN_LW
    ws.merge_cells(start_row=cur_row, end_row=cur_row+1, start_column=1, end_column=8)
    cur_row += 2

    # ── 5. ACTION BAR ────────────────────────────────────────────────────
    cur_row += 1
    paint_row(cur_row, COL_BG_CARD, height=8)
    cur_row += 1

    action_cells = [
        ("ACTION", action.get("entry_label", "—"), COL_MINT_400, COL_BG_ACTION),
        ("STOP LOSS",
            f"₹{action.get('stop_loss_rs', 0):,.2f}  ({action.get('stop_loss_pct', -7)}%)",
            COL_WHITE_85, COL_BG_STAT),
        ("TIME HORIZON",
            f"Up to {action.get('time_horizon_days', 20)} trading days",
            COL_WHITE_85, COL_BG_STAT),
    ]
    paint_row(cur_row,   COL_BG_CARD, height=18)
    paint_row(cur_row+1, COL_BG_CARD, height=24)
    paint_row(cur_row+2, COL_BG_CARD, height=8)
    col_groups = [(1, 3), (4, 5), (6, 8)]   # 3-2-3 column split = 8 wide
    for i, (label, val, color, bg) in enumerate(action_cells):
        cs, ce = col_groups[i]
        for c in range(cs, ce + 1):
            for rr in (cur_row, cur_row+1):
                ws.cell(row=rr, column=c).fill = FILL(bg)
                ws.cell(row=rr, column=c).border = BORDER(COL_BORDER_MINT if bg == COL_BG_ACTION else COL_BORDER)
        ws.merge_cells(start_row=cur_row, end_row=cur_row, start_column=cs, end_column=ce)
        ws.merge_cells(start_row=cur_row+1, end_row=cur_row+1, start_column=cs, end_column=ce)
        lc = ws.cell(row=cur_row, column=cs, value=label)
        lc.font = F(size=8, bold=True, color=COL_WHITE_45); lc.alignment = ALN_L
        vc = ws.cell(row=cur_row+1, column=cs, value=val)
        vc.font = F(size=11, bold=True, color=color); vc.alignment = ALN_L

    # Hide gridlines on this sheet
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 110

    return name


def _draw_stat_row(ws, row: int, stats: list[tuple]) -> None:
    """Draws 4 stat boxes (2 cols × 3 rows each) starting at the given row."""
    col_groups = [(1, 2), (3, 4), (5, 6), (7, 8)]
    for r in (row, row + 1, row + 2):
        ws.row_dimensions[r].height = 18 if r != row + 1 else 26
        for c in range(1, 9):
            ws.cell(row=r, column=c).fill = FILL(COL_BG_STAT)
            ws.cell(row=r, column=c).border = BORDER()
    for i, (label, value, hint, value_color) in enumerate(stats):
        cs, ce = col_groups[i]
        ws.merge_cells(start_row=row,     end_row=row,     start_column=cs, end_column=ce)
        ws.merge_cells(start_row=row + 1, end_row=row + 1, start_column=cs, end_column=ce)
        ws.merge_cells(start_row=row + 2, end_row=row + 2, start_column=cs, end_column=ce)
        ws.cell(row=row,     column=cs, value=label).font = F(size=8, bold=True, color=COL_WHITE_45)
        ws.cell(row=row,     column=cs).alignment = ALN_L
        vc = ws.cell(row=row + 1, column=cs, value=value)
        vc.font = F(size=16, bold=True, color=value_color, name="Consolas")
        vc.alignment = ALN_L
        ws.cell(row=row + 2, column=cs, value=hint).font = F(size=8, italic=True, color=COL_WHITE_45)
        ws.cell(row=row + 2, column=cs).alignment = ALN_L


def _bucket2_callout_text(b2: dict) -> str:
    nout = b2.get("n_with_outcome", 0)
    if nout > 0:
        hr     = b2.get("hit_rate_pct", 0)
        base   = b2.get("lifetime_baseline_pct", 0)
        delta  = b2.get("kaynes_delta_pp", 0)
        dir_word = "stronger" if delta >= 0 else "weaker"
        sign = "+" if delta >= 0 else ""
        text = (f"Today's pattern set has historically delivered a {hr:.0f}% +10%/20d hit rate "
                f"on this stock — that's {sign}{delta:.0f}pp {dir_word} than the stock's "
                f"lifetime baseline ({base:.0f}%).")
        last = b2.get("last_similar")
        if last:
            text += (f" Most recent similar setup fired {last['date']} → "
                     f"{last['return_pct']:+.1f}% in {last['days_to_outcome']} days.")
        return text
    n_hist = b2.get("n_historical_fires", 0)
    base   = b2.get("lifetime_baseline_pct", 0)
    lb     = b2.get("lookback_start", "2021-01-01")
    if n_hist > 0:
        return (f"{n_hist:,} historical pattern matches on this stock since {lb}, but "
                f"they're too recent for 20-day outcomes to have realised. The stock's "
                f"lifetime baseline above is the honest reference: about {base:.0f}% of "
                f"any random entry on this stock hits +10% in 20 trading days.")
    return (f"The patterns fired today have never historically matched this stock's feature "
            f"signature in our backtested window since {lb}. The stock's lifetime baseline "
            f"+10%/20d rate is {base:.0f}%.")


def _sector_narrative(b3: dict) -> str:
    sec = b3.get("sector_name", "the sector")
    rank = b3.get("sector_rank", 0)
    total = b3.get("total_sectors", 0)
    peers = b3.get("n_peers_firing", 0)
    rot = b3.get("rotation_direction", "neutral")
    sessions = b3.get("rotation_sessions_of_10", 5)

    if rank == 0:
        rankp = f"is unranked today (no other names firing in this sector)"
    elif rank <= 3:
        rankp = f"is leading the market today (rank #{rank} of {total})"
    elif rank <= 7:
        rankp = f"is mid-pack today (rank #{rank} of {total})"
    else:
        rankp = f"is lagging today (rank #{rank} of {total})"

    if peers >= 3:
        peersp = f"with {peers} other names in the sector also firing patterns — broad-based strength"
    elif peers >= 1:
        peersp = f"with {peers} other name{'s' if peers > 1 else ''} firing alongside it"
    else:
        peersp = "as the only name in its sector firing today — idiosyncratic move, less sector tailwind"

    if rot == "inflow":
        rotp = f" Capital has been rotating into the sector {sessions}/10 recent sessions."
    elif rot == "outflow":
        rotp = f" Capital has been rotating OUT of the sector {sessions}/10 recent sessions — headwind."
    else:
        rotp = ""

    return f"The {sec} sector {rankp}, {peersp}.{rotp}"


def _verdict_color(verdict: str | None) -> str:
    return {
        "tailwind": COL_MINT_400,
        "headwind": COL_RED_300,
        "neutral":  COL_AMBER_300,
        "unranked": COL_WHITE_55,
    }.get(verdict or "neutral", COL_WHITE_55)


REGIME_HUMAN = {
    "breakout":       "Breakout",
    "momentum":       "Momentum",
    "compression":    "Compression",
    "reversal":       "Reversal",
    "mean_reversion": "Mean-reversion",
    "capitulation":   "Capitulation",
    "other":          "Multi-factor",
}
REGIME_GLYPH = {
    "breakout":       "★",
    "momentum":       "↗",
    "compression":    "⟳",
    "reversal":       "↺",
    "mean_reversion": "⇋",
    "capitulation":   "⚠",
    "other":          "·",
}
REGIME_COLOR = {
    "breakout":       COL_MINT_400,
    "momentum":       COL_BLUE_300,
    "compression":    COL_AMBER_300,
    "reversal":       "C4B5FD",
    "mean_reversion": "C4B5FD",
    "capitulation":   COL_RED_300,
    "other":          COL_WHITE_55,
}


# ── Index sheet ─────────────────────────────────────────────────────────
def write_index(wb: Workbook, payload: dict, sheet_names: list[str]) -> None:
    ws = wb.create_sheet("Index", 0)
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 8, "B": 14, "C": 26, "D": 16, "E": 12, "F": 12, "G": 12, "H": 22}.items():
        ws.column_dimensions[col].width = w

    # Title strip
    for c in range(1, 9):
        ws.cell(row=1, column=c).fill = FILL(COL_BG_CARD)
        ws.cell(row=2, column=c).fill = FILL(COL_BG_CARD)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18

    tcell = ws.cell(row=1, column=1,
                     value=f"KANIDA.AI  ·  Falcon Top 10  ·  {payload.get('signal_date', '—')}")
    tcell.font = F(size=16, bold=True, color=COL_MINT_400)
    tcell.alignment = ALN_L
    ws.merge_cells("A1:H1")
    sub = ws.cell(row=2, column=1,
                   value=(f"Signal {payload.get('signal_date', '—')}  ·  "
                          f"Entry {payload.get('entry_date', '—')}  ·  "
                          f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M IST')}  ·  "
                          f"Click any row to jump to that pick's card"))
    sub.font = F(size=9, italic=True, color=COL_WHITE_55); sub.alignment = ALN_L
    ws.merge_cells("A2:H2")

    # Header row
    hdrs = ["#", "Symbol", "Sector", "Signal Type", "Risk", "Today HR", "Flags", "Action"]
    for c in range(1, 9):
        ws.cell(row=4, column=c).fill = FILL(COL_BG_SECTION)
        ws.cell(row=4, column=c).border = BORDER()
    ws.row_dimensions[4].height = 22
    for i, h in enumerate(hdrs, start=1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = F(size=10, bold=True, color=COL_WHITE_85)
        cell.alignment = ALN_C if i in (1, 5, 6) else ALN_L

    # One row per pick
    for i, (p, sname) in enumerate(zip(payload["picks"], sheet_names), start=5):
        b1 = p.get("bucket1", {}) or {}
        f_ = p.get("flags", {}) or {}
        act = p.get("action", {}) or {}
        thr = b1.get("today_weighted_hit_rate")
        thr_str = f"{round(thr)}%" if thr is not None else "N/A"
        flag_list = []
        if f_.get("fresh_entry"):         flag_list.append("FRESH")
        if f_.get("falling_knife"):       flag_list.append("⚠ KNIFE")
        elif f_.get("capitulation_warning"): flag_list.append("⚠ CAP")

        for c in range(1, 9):
            ws.cell(row=i, column=c).fill = FILL(COL_BG_CARD)
            ws.cell(row=i, column=c).border = BORDER()
        ws.row_dimensions[i].height = 20

        rank_cell = ws.cell(row=i, column=1, value=f"#{p['rank']}")
        rank_cell.font = F(size=11, bold=True, color=COL_MINT_400, name="Consolas")
        rank_cell.alignment = ALN_C

        sym_cell = ws.cell(row=i, column=2, value=p["symbol"])
        sym_cell.font = F(size=11, bold=True, color=COL_WHITE)
        sym_cell.alignment = ALN_L
        # Internal hyperlink to the stock sheet
        # Quote the sheet name if it has spaces
        link_target = f"'{sname}'!A1"
        sym_cell.hyperlink = Hyperlink(ref=sym_cell.coordinate, location=link_target,
                                        display=p["symbol"], tooltip=f"Jump to {p['symbol']}")
        sym_cell.font = F(size=11, bold=True, color=COL_MINT_400, name="Calibri")

        ws.cell(row=i, column=3, value=p.get("sector") or "—").font = F(color=COL_WHITE_85)
        ws.cell(row=i, column=3).alignment = ALN_L

        st = ws.cell(row=i, column=4, value=p.get("signal_type", "—"))
        st.font = F(color=REGIME_COLOR.get(p.get("signal_type", "").lower(), COL_WHITE_85))
        st.alignment = ALN_L

        rk = ws.cell(row=i, column=5, value=p.get("risk_level", "—"))
        rk_col = (COL_MINT_300 if p.get("risk_level") == "Low"
                  else COL_RED_300 if p.get("risk_level") == "High"
                  else COL_AMBER_300)
        rk.font = F(color=rk_col, bold=True); rk.alignment = ALN_C

        hr_c = ws.cell(row=i, column=6, value=thr_str)
        hr_c.font = F(color=COL_MINT_300, name="Consolas", bold=True); hr_c.alignment = ALN_C

        ws.cell(row=i, column=7, value=", ".join(flag_list) if flag_list else "—").font = F(
            color=(COL_MINT_400 if "FRESH" in flag_list and "KNIFE" not in flag_list
                   else COL_RED_300 if "KNIFE" in flag_list
                   else COL_AMBER_300 if "CAP" in flag_list
                   else COL_WHITE_55))

        ws.cell(row=i, column=8, value=act.get("entry_label", "—")).font = F(color=COL_WHITE_55, size=9)


# ── Driver ──────────────────────────────────────────────────────────────
def main() -> None:
    payload = fetch_payload()
    wb = Workbook()
    wb.remove(wb.active)

    sheet_names: list[str] = []
    for p in payload.get("picks", []):
        name = write_stock_sheet(wb, p, payload.get("signal_date", ""),
                                  payload.get("entry_date", ""))
        sheet_names.append(name)

    write_index(wb, payload, sheet_names)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    size_kb = OUT_PATH.stat().st_size / 1024
    print(f"\n[OK] wrote {OUT_PATH}  ({size_kb:.1f} KB)")
    print(f"     {len(sheet_names) + 1} sheets: Index + {len(sheet_names)} per-stock cards")


if __name__ == "__main__":
    main()
